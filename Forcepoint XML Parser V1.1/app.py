"""
Forcepoint XML Parser — Web Viewer + Excel Exporter
pip install flask openpyxl
python app_v3.py  ->  http://localhost:5000
"""

from flask import Flask, request, jsonify, send_file, render_template_string
import xml.etree.ElementTree as ET
import re, ipaddress, io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ═══════════════════════════════════════════════════════════
# NAME SANITIZER
# ═══════════════════════════════════════════════════════════
_name_map = {}

def sanitize_name(orig: str, max_len=79) -> str:
    s = orig.strip()
    s = re.sub(r'[^\w\-\.]', '_', s)
    s = re.sub(r'_+', '_', s)
    s = s.strip('_').rstrip('.')
    s = s or 'obj'
    s = s[:max_len]
    if s != orig:
        _name_map[orig] = s
    return s


# ═══════════════════════════════════════════════════════════
# FORCEPOINT PARSER
# ═══════════════════════════════════════════════════════════
def extract_objects(xml_bytes):
    global _name_map
    _name_map = {}
    root = ET.fromstring(xml_bytes)

    xml_object_tags = ('host', 'network', 'address_range', 'group',
                       'gen_service_group', 'service_tcp', 'service_udp',
                       'interface', 'zone', 'fw_policy', 'access_rule')
    xml_names = set()
    for tag in xml_object_tags:
        for el in root.iter(tag):
            n = el.attrib.get('name', '').strip()
            if n:
                xml_names.add(n)
                sanitize_name(n)

    def resolve(ref):
        if ref in xml_names:
            return _name_map.get(ref, ref)
        return ref

    # ── 1. Hosts ──────────────────────────────────────────────
    hosts = []
    for el in root.iter('host'):
        ip_el = el.find('mvia_address')
        ip = ip_el.attrib.get('address', '').strip() if ip_el is not None else ''
        if not ip:
            ip = el.attrib.get('address', '').strip()
        orig = el.attrib.get('name', '').strip()
        if not orig: continue
        hosts.append({"name": _name_map.get(orig, orig), "ip": ip})

    # ── 2. Networks ────────────────────────────────────────────
    networks = []
    for el in root.iter('network'):
        net  = el.attrib.get('ipv4_network', '').strip()
        orig = el.attrib.get('name', '').strip()
        if not orig: continue
        ip, mask = '', ''
        if '/' in net:
            try:
                n = ipaddress.IPv4Network(net, strict=False)
                ip, mask = str(n.network_address), str(n.netmask)
            except: pass
        networks.append({"name": _name_map.get(orig, orig), "ip": ip, "mask": mask})

    # ── 3. Address Ranges ──────────────────────────────────────
    addr_ranges = []
    for el in root.iter('address_range'):
        r    = el.attrib.get('ip_range', '').strip()
        orig = el.attrib.get('name', '').strip()
        if not orig: continue
        addr_ranges.append({"name": _name_map.get(orig, orig), "range": r})

    # ── 4. Address Groups ──────────────────────────────────────
    net_groups = []
    for el in root.iter('group'):
        orig = el.attrib.get('name', '').strip()
        if not orig: continue
        members = []
        for ne in el.findall('ne_list'):
            ref = ne.attrib.get('ref', '').strip()
            if ref: members.append(resolve(ref))
        seen = set()
        members = [m for m in members if not (m in seen or seen.add(m))]
        net_groups.append({"name": _name_map.get(orig, orig), "members": ', '.join(members)})

    # ── 5. Services ────────────────────────────────────────────
    services = []
    for el in root.iter('service_tcp'):
        min_p = el.attrib.get('min_dst_port', '').strip()
        max_p = el.attrib.get('max_dst_port', min_p).strip()
        orig  = el.attrib.get('name', '').strip()
        if not orig or not min_p: continue
        val = f"{min_p}-{max_p}" if max_p and max_p != min_p else min_p
        services.append({"name": _name_map.get(orig, orig), "value": val, "type": "TCP"})
    for el in root.iter('service_udp'):
        min_p = el.attrib.get('min_dst_port', '').strip()
        max_p = el.attrib.get('max_dst_port', min_p).strip()
        orig  = el.attrib.get('name', '').strip()
        if not orig or not min_p: continue
        val = f"{min_p}-{max_p}" if max_p and max_p != min_p else min_p
        services.append({"name": _name_map.get(orig, orig), "value": val, "type": "UDP"})
    for el in root.iter('service_icmp'):
        orig = el.attrib.get('name', '').strip()
        if not orig: continue
        services.append({"name": _name_map.get(orig, orig), "value": "ICMP", "type": "ICMP"})

    # ── 6. Service Groups ──────────────────────────────────────
    service_groups = []
    for el in root.iter('gen_service_group'):
        orig = el.attrib.get('name', '').strip()
        if not orig: continue
        members = []
        for s in el.findall('service_ref'):
            ref = s.attrib.get('ref', '').strip()
            if ref: members.append(resolve(ref))
        seen = set()
        members = [m for m in members if not (m in seen or seen.add(m))]
        service_groups.append({"name": _name_map.get(orig, orig), "members": ', '.join(members)})

    # ── 7. Zones ───────────────────────────────────────────────
    zones = []
    seen_zones = set()
    for tag in ('zone', 'interface_zone', 'fw_zone'):
        for el in root.iter(tag):
            orig = el.attrib.get('name', '').strip()
            if orig and orig not in seen_zones:
                seen_zones.add(orig)
                zones.append({"name": _name_map.get(orig, orig)})
    for tag in ('access_rule', 'fw_policy', 'rule'):
        for el in root.iter(tag):
            for attr in ('src_zone_ref', 'dst_zone_ref', 'source_zone', 'destination_zone', 'from_zone', 'to_zone'):
                ref = el.attrib.get(attr, '').strip()
                if ref and ref not in seen_zones:
                    seen_zones.add(ref)
                    zones.append({"name": resolve(ref)})

    # ── 8. Policies ────────────────────────────────────────────
    # Real Forcepoint XML structure:
    # fw_policy > access_entry > rule_entry > access_rule > match_part
    #   match_part > match_sources    > match_source_ref      [value=name]
    #             > match_destinations> match_destination_ref  [value=name]
    #             > match_services    > match_service_ref      [value=name]
    # access_rule > action            [type=allow/discard/...]
    # access_rule > option > log_policy [log_level=stored/undefined/...]
    # rule_entry attribs: name (optional), is_disabled, comment
    policies = []
    rule_counter = 0

    def _collect_refs(parent_el, child_tag):
        """Collect all value= from child elements of parent_el with tag child_tag."""
        el = parent_el.find(child_tag)
        if el is None:
            return ''
        vals = []
        seen = set()
        for ref in el:
            v = ref.attrib.get('value', '').strip()
            if v and v not in seen:
                seen.add(v)
                vals.append(v)
        return ', '.join(vals)

    for fp in root.iter('fw_policy'):
        for re_el in fp.iter('rule_entry'):
            ar = re_el.find('access_rule')
            if ar is None:
                continue

            rule_counter += 1
            # Name: rule_entry may have 'name', else use comment, else auto
            name = re_el.attrib.get('name', '').strip()
            if not name:
                comment = re_el.attrib.get('comment', '').strip()
                name = comment[:60] if comment else f"Rule_{rule_counter}"

            # Status
            is_disabled = re_el.attrib.get('is_disabled', 'false').lower()
            status = 'Disabled' if is_disabled in ('true', '1') else 'Enabled'

            # match_part holds sources / destinations / services
            mp = ar.find('match_part')
            src_net  = _collect_refs(mp, 'match_sources')      if mp is not None else ''
            dst_net  = _collect_refs(mp, 'match_destinations') if mp is not None else ''
            rule_svcs = _collect_refs(mp, 'match_services')     if mp is not None else ''

            # Action (allow / discard / refuse / jump / continue)
            action_el = ar.find('action')
            action = action_el.attrib.get('type', '') if action_el is not None else ''

            # Log level
            log_val = ''
            opt = ar.find('option')
            if opt is not None:
                lp = opt.find('log_policy')
                if lp is not None:
                    log_val = lp.attrib.get('log_level', '')

            # Forcepoint policies don't have separate src/dst zone per rule
            # (zones are at the engine/interface level, not the rule level)
            src_zone = ''
            dst_zone = ''
            src_port = ''   # Forcepoint combines src/dst services in match_services
            dst_port = rule_svcs

            policies.append({
                "name":     name,
                "src_zone": src_zone,
                "dst_zone": dst_zone,
                "src_net":  src_net,
                "dst_net":  dst_net,
                "src_port": src_port,
                "dst_port": dst_port,
                "log":      log_val,
                "status":   status,
                "action":   action,
            })

    return {
        "hosts": hosts, "networks": networks, "addr_ranges": addr_ranges,
        "net_groups": net_groups, "services": services, "service_groups": service_groups,
        "zones": zones, "policies": policies,
    }


# ═══════════════════════════════════════════════════════════
# EXCEL BUILDER
# ═══════════════════════════════════════════════════════════
C_HEADER_BG = "1A2D45"
C_HEADER_FG = "E8F4FD"
C_ALT_ROW   = "F0F8FF"
C_BORDER    = "B8D4E8"
C_TAB_COLORS = {
    "Hosts":          "2196F3", "Networks":       "4CAF50",
    "Address_Ranges": "FF9800", "Address_Groups": "9C27B0",
    "Services":       "F44336", "Service_Groups": "E91E63",
    "Zones":          "00BCD4", "Policies":       "607D8B",
}

def _hdr(ws, row, cols, tab_color):
    fill = PatternFill("solid", fgColor=C_HEADER_BG)
    font = Font(bold=True, color=C_HEADER_FG, name="Arial", size=10)
    aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=C_BORDER)
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.font = font; c.fill = fill; c.alignment = aln; c.border = brd
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row+1, column=1)
    ws.sheet_properties.tabColor = tab_color

def _row(ws, row, vals, alt=False):
    fill = PatternFill("solid", fgColor=C_ALT_ROW) if alt else None
    thin = Side(style="thin", color=C_BORDER)
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    aln  = Alignment(vertical="center", wrap_text=True)
    font = Font(name="Arial", size=9)
    for i, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.border = brd; c.alignment = aln; c.font = font
        if fill: c.fill = fill
    ws.row_dimensions[row].height = 16

def _title(ws, title, ncols, color):
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, name="Arial", size=13, color="07111A")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_excel(data):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Hosts")
    _hdr(ws, 1, ["Name", "IP Address"], C_TAB_COLORS["Hosts"])
    for i, r in enumerate(data["hosts"]):
        _row(ws, i+2, [r["name"], r["ip"]], i%2==1)
    _widths(ws, [35, 20])
    _title(ws, f"Hosts  ({len(data['hosts'])} records)", 2, C_TAB_COLORS["Hosts"])

    ws = wb.create_sheet("Networks")
    _hdr(ws, 1, ["Name", "Network IP", "Subnet Mask"], C_TAB_COLORS["Networks"])
    for i, r in enumerate(data["networks"]):
        _row(ws, i+2, [r["name"], r["ip"], r["mask"]], i%2==1)
    _widths(ws, [35, 18, 18])
    _title(ws, f"Networks  ({len(data['networks'])} records)", 3, C_TAB_COLORS["Networks"])

    ws = wb.create_sheet("Address_Ranges")
    _hdr(ws, 1, ["Name", "IP Range"], C_TAB_COLORS["Address_Ranges"])
    for i, r in enumerate(data["addr_ranges"]):
        _row(ws, i+2, [r["name"], r["range"]], i%2==1)
    _widths(ws, [35, 35])
    _title(ws, f"Address Ranges  ({len(data['addr_ranges'])} records)", 2, C_TAB_COLORS["Address_Ranges"])

    ws = wb.create_sheet("Address_Groups")
    _hdr(ws, 1, ["Name", "Members"], C_TAB_COLORS["Address_Groups"])
    for i, r in enumerate(data["net_groups"]):
        _row(ws, i+2, [r["name"], r["members"]], i%2==1)
    _widths(ws, [35, 80])
    _title(ws, f"Address Groups  ({len(data['net_groups'])} records)", 2, C_TAB_COLORS["Address_Groups"])

    ws = wb.create_sheet("Services")
    _hdr(ws, 1, ["Name", "Value / Port", "Type"], C_TAB_COLORS["Services"])
    for i, r in enumerate(data["services"]):
        _row(ws, i+2, [r["name"], r["value"], r["type"]], i%2==1)
    _widths(ws, [35, 20, 10])
    _title(ws, f"Services  ({len(data['services'])} records)", 3, C_TAB_COLORS["Services"])

    ws = wb.create_sheet("Service_Groups")
    _hdr(ws, 1, ["Name", "Members"], C_TAB_COLORS["Service_Groups"])
    for i, r in enumerate(data["service_groups"]):
        _row(ws, i+2, [r["name"], r["members"]], i%2==1)
    _widths(ws, [35, 80])
    _title(ws, f"Service Groups  ({len(data['service_groups'])} records)", 2, C_TAB_COLORS["Service_Groups"])

    ws = wb.create_sheet("Zones")
    _hdr(ws, 1, ["Name"], C_TAB_COLORS["Zones"])
    for i, r in enumerate(data["zones"]):
        _row(ws, i+2, [r["name"]], i%2==1)
    _widths(ws, [40])
    _title(ws, f"Zones  ({len(data['zones'])} records)", 1, C_TAB_COLORS["Zones"])

    ws = wb.create_sheet("Policies")
    _hdr(ws, 1, ["Name","Source Network","Destination Network","Service","Action","Log","Status"],
         C_TAB_COLORS["Policies"])
    for i, r in enumerate(data["policies"]):
        _row(ws, i+2, [r["name"], r["src_net"], r["dst_net"],
                       r["dst_port"], r.get("action",""), r["log"], r["status"]], i%2==1)
    _widths(ws, [40, 40, 40, 35, 12, 12, 12])
    _title(ws, f"Policies  ({len(data['policies'])} records)", 7, C_TAB_COLORS["Policies"])

    # Summary
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "263238"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "Forcepoint Export Summary"
    c.font  = Font(bold=True, size=14, name="Arial", color="E8F4FD")
    c.fill  = PatternFill("solid", fgColor="0D1829")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    thin = Side(style="thin", color=C_BORDER)
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    rows_data = [("Sheet","Records"),("Hosts",len(data["hosts"])),("Networks",len(data["networks"])),
                 ("Address Ranges",len(data["addr_ranges"])),("Address Groups",len(data["net_groups"])),
                 ("Services",len(data["services"])),("Service Groups",len(data["service_groups"])),
                 ("Zones",len(data["zones"])),("Policies",len(data["policies"]))]
    for i, (label, val) in enumerate(rows_data):
        r = i+2
        ca = ws.cell(row=r, column=1, value=label)
        cb = ws.cell(row=r, column=2, value=val)
        for c in (ca, cb):
            c.border = brd
            c.alignment = Alignment(horizontal="center", vertical="center")
            if i == 0:
                c.font = Font(bold=True, name="Arial", size=10, color=C_HEADER_FG)
                c.fill = PatternFill("solid", fgColor=C_HEADER_BG)
            else:
                c.font = Font(name="Arial", size=10)
                if i%2==0: c.fill = PatternFill("solid", fgColor=C_ALT_ROW)
        ws.row_dimensions[r].height = 22
    r = len(rows_data)+2
    total = sum(len(data[k]) for k in data)
    for col, val in [(1,"TOTAL"),(2,total)]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(bold=True, name="Arial", size=11, color="07111A")
        c.fill = PatternFill("solid", fgColor="00C8F0")
        c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
# HTML
# ═══════════════════════════════════════════════════════════
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Forcepoint XML Parser</title>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=Inter:wght@400;600;700;900&display=swap" rel="stylesheet"/>
<style>
:root{
  --bg:#070f1a;--sur:#0d1829;--card:#111f33;--bdr:#1a2d45;
  --acc:#00c8f0;--acc2:#6d28d9;--grn:#10b981;--org:#f59e0b;
  --red:#ef4444;--txt:#dde6f0;--mut:#4e6680;
  --mono:'IBM Plex Mono',monospace;--sans:'Inter',sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg);color:var(--txt);font-family:var(--sans);min-height:100vh}
body::before{content:'';position:fixed;inset:0;
  background-image:linear-gradient(rgba(0,200,240,.018) 1px,transparent 1px),
                   linear-gradient(90deg,rgba(0,200,240,.018) 1px,transparent 1px);
  background-size:48px 48px;pointer-events:none;z-index:0}
.wrap{position:relative;z-index:1;max-width:1700px;margin:0 auto;padding:0 24px 60px}

/* ── Header ── */
header{padding:22px 0 18px;border-bottom:1px solid var(--bdr);display:flex;align-items:center;gap:14px;margin-bottom:28px}
.logo{width:44px;height:44px;background:linear-gradient(135deg,var(--acc),var(--acc2));border-radius:11px;
  display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0}
.brand h1{font-size:19px;font-weight:900;letter-spacing:-.3px}
.brand p{font-size:11px;color:var(--mut);font-family:var(--mono);margin-top:2px}
#hdr-right{margin-left:auto;display:flex;gap:10px;align-items:center}

/* ── Buttons ── */
.btn{display:inline-flex;align-items:center;gap:7px;padding:9px 20px;border-radius:8px;
  border:none;font-family:var(--sans);font-size:13px;font-weight:600;cursor:pointer;transition:all .2s}
.btn:disabled{opacity:.4;cursor:not-allowed}
.btn-sm{padding:6px 14px;font-size:12px;border-radius:7px}
.btn-primary{background:linear-gradient(135deg,var(--acc),var(--acc2));color:#fff}
.btn-primary:not(:disabled):hover{opacity:.87;transform:translateY(-1px)}
.btn-outline{background:transparent;border:1px solid var(--bdr);color:var(--txt)}
.btn-outline:not(:disabled):hover{border-color:var(--acc);color:var(--acc)}
.btn-green{background:var(--grn);color:#fff}.btn-green:not(:disabled):hover{opacity:.85}

/* ── Upload ── */
#upload-section{padding:40px 0;text-align:center}
.drop-zone{border:2px dashed var(--bdr);border-radius:18px;padding:60px 40px;
  cursor:pointer;transition:all .3s;background:var(--sur);position:relative;overflow:hidden;max-width:600px;margin:0 auto}
.drop-zone::before{content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse at center,rgba(0,200,240,.05) 0%,transparent 70%);
  opacity:0;transition:opacity .3s}
.drop-zone:hover,.drop-zone.drag{border-color:var(--acc)}
.drop-zone:hover::before,.drop-zone.drag::before{opacity:1}
#file-input{display:none}

/* ── Progress ── */
#progress-section{display:none;text-align:center;padding:60px 0}
.pbar-wrap{height:5px;background:var(--bdr);border-radius:99px;overflow:hidden;margin:16px auto;max-width:400px}
.pbar{height:100%;background:linear-gradient(90deg,var(--acc),var(--acc2));border-radius:99px;width:0%;transition:width .4s}
#plabel{font-family:var(--mono);font-size:11px;color:var(--mut);margin-top:8px}

/* ── Results ── */
#results-section{display:none}

/* Summary strip */
.summary-strip{display:flex;gap:10px;margin-bottom:22px;flex-wrap:wrap}
.scard{background:var(--sur);border:1px solid var(--bdr);border-radius:10px;
  padding:12px 18px;display:flex;align-items:center;gap:12px;cursor:pointer;transition:all .2s;flex:1;min-width:130px}
.scard:hover{border-color:var(--acc)}
.scard.active{border-color:var(--acc);background:rgba(0,200,240,.08)}
.scard-dot{width:10px;height:10px;border-radius:50%;flex-shrink:0}
.scard-info .val{font-size:20px;font-weight:900;font-family:var(--mono)}
.scard-info .lbl{font-size:10px;color:var(--mut);text-transform:uppercase;letter-spacing:.4px;margin-top:1px}

/* Toolbar */
.toolbar{display:flex;gap:8px;align-items:center;margin-bottom:14px;flex-wrap:wrap}
.search-box{background:var(--sur);border:1px solid var(--bdr);border-radius:8px;
  padding:8px 13px;color:var(--txt);font-family:var(--mono);font-size:12px;outline:none;
  transition:border-color .2s;width:240px}
.search-box:focus{border-color:var(--acc)}
.search-box::placeholder{color:var(--mut)}
.sep{width:1px;height:24px;background:var(--bdr)}
.tab-label{font-family:var(--mono);font-size:12px;font-weight:600;padding:6px 14px;
  border-radius:6px;border:1px solid transparent;cursor:pointer;transition:all .2s;color:var(--mut)}
.tab-label.active{color:#fff;border-color:transparent}

/* Table */
.table-wrap{background:var(--sur);border:1px solid var(--bdr);border-radius:13px;overflow:hidden}
.table-scroll{overflow:auto;max-height:65vh}
.table-scroll::-webkit-scrollbar{width:5px;height:5px}
.table-scroll::-webkit-scrollbar-thumb{background:var(--bdr);border-radius:99px}
table{width:100%;border-collapse:collapse;font-family:var(--mono);font-size:11.5px}
thead th{position:sticky;top:0;background:#0a1525;color:var(--acc);padding:10px 14px;
  text-align:left;font-weight:600;white-space:nowrap;border-bottom:1px solid var(--bdr);
  cursor:pointer;user-select:none}
thead th:hover{color:#fff}
thead th .sort-ico{opacity:.4;font-size:9px;margin-left:4px}
thead th.asc .sort-ico::after{content:'▲'}
thead th.desc .sort-ico::after{content:'▼'}
thead th:not(.asc):not(.desc) .sort-ico::after{content:'⇅'}
tbody tr{border-bottom:1px solid rgba(255,255,255,.025)}
tbody tr:hover{background:rgba(0,200,240,.04)}
tbody td{padding:8px 14px;vertical-align:top}
tbody td.wrap-cell{max-width:320px;white-space:pre-wrap;word-break:break-word;line-height:1.5}
tbody td.nowrap{white-space:nowrap}
.badge{display:inline-block;padding:2px 8px;border-radius:99px;font-size:10px;font-weight:700}
.badge-tcp{background:rgba(33,150,243,.2);color:#64b5f6}
.badge-udp{background:rgba(255,152,0,.2);color:#ffb74d}
.badge-icmp{background:rgba(76,175,80,.2);color:#81c784}
.badge-en{background:rgba(16,185,129,.15);color:#10b981}
.badge-dis{background:rgba(239,68,68,.15);color:#ef4444}

/* Pagination */
.pagination{padding:10px 18px;border-top:1px solid var(--bdr);display:flex;align-items:center;gap:6px}
.pg-info{font-family:var(--mono);font-size:10px;color:var(--mut);flex:1}
.pg-btn{background:var(--bdr);border:none;color:var(--txt);padding:4px 10px;border-radius:5px;
  cursor:pointer;font-family:var(--mono);font-size:11px;transition:background .15s}
.pg-btn:hover:not(:disabled){background:var(--acc);color:#000}
.pg-btn:disabled{opacity:.35;cursor:not-allowed}
.pg-btn.active{background:var(--acc);color:#000}

/* Empty */
.empty-state{padding:60px;text-align:center;color:var(--mut);font-family:var(--mono)}

/* Toast */
.toast{position:fixed;bottom:20px;left:50%;transform:translateX(-50%);background:var(--grn);
  color:#fff;padding:10px 22px;border-radius:8px;font-weight:700;font-size:12px;
  opacity:0;transition:opacity .3s;z-index:200;pointer-events:none;white-space:nowrap}
.toast.show{opacity:1}

@media(max-width:900px){.summary-strip{flex-wrap:wrap}}
</style>
</head>
<body>
<div class="wrap">

<!-- Header -->
<header>
  <div class="logo">🛡️</div>
  <div class="brand">
    <h1>Forcepoint XML Parser</h1>
    <p>Created By Genius</p>
  </div>
  <div id="hdr-right"></div>
</header>

<!-- Upload -->
<div id="upload-section">
  <div class="drop-zone" id="drop-zone" onclick="document.getElementById('file-input').click()">
    <div style="font-size:52px;margin-bottom:12px">📂</div>
    <div style="font-size:20px;font-weight:700;margin-bottom:8px">Drop your Forcepoint XML here</div>
    <div style="color:var(--mut);font-size:12px;font-family:var(--mono);margin-bottom:22px">drag &amp; drop or click to browse .xml files only</div>
    <button class="btn btn-primary" onclick="event.stopPropagation();document.getElementById('file-input').click()">📁 Browse File</button>
  </div>
  <input type="file" id="file-input" accept=".xml" onchange="handleFile(this.files[0])"/>
</div>

<!-- Progress -->
<div id="progress-section">
  <div style="font-size:46px;margin-bottom:12px">⚙️</div>
  <div style="font-size:17px;font-weight:700;margin-bottom:14px">Parsing XML...</div>
  <div class="pbar-wrap"><div class="pbar" id="pbar"></div></div>
  <div id="plabel">reading...</div>
</div>

<!-- Results -->
<div id="results-section">
  <!-- Summary cards -->
  <div class="summary-strip" id="summary-strip"></div>

  <!-- Toolbar -->
  <div class="toolbar">
    <input class="search-box" id="search-box" placeholder="🔍 Search in table..." oninput="onSearch()"/>
    <div class="sep"></div>
    <button class="btn btn-sm btn-outline" onclick="exportCSV()">⬇ CSV</button>
    <button class="btn btn-sm btn-green" onclick="downloadExcel()" id="btn-xl">⬇ Excel (All Sheets)</button>
    <button class="btn btn-sm btn-outline" onclick="resetApp()">↺ New File</button>
  </div>

  <!-- Table -->
  <div class="table-wrap">
    <div class="table-scroll">
      <table>
        <thead id="thead"></thead>
        <tbody id="tbody"></tbody>
      </table>
    </div>
    <div class="pagination">
      <div class="pg-info" id="pg-info"></div>
      <button class="pg-btn" id="btn-prev" onclick="changePage(-1)">«</button>
      <div id="pg-nums"></div>
      <button class="pg-btn" id="btn-next" onclick="changePage(1)">»</button>
    </div>
  </div>
</div>

</div>
<div class="toast" id="toast"></div>

<script>
// ═══ Config ══════════════════════════════════════════════
const SHEETS = [
  { key:'hosts',          label:'Hosts',           color:'#2196F3',
    cols:['Name','IP Address'],
    row: r => [r.name, r.ip] },
  { key:'networks',       label:'Networks',         color:'#4CAF50',
    cols:['Name','Network IP','Subnet Mask'],
    row: r => [r.name, r.ip, r.mask] },
  { key:'addr_ranges',    label:'Address Ranges',   color:'#FF9800',
    cols:['Name','IP Range'],
    row: r => [r.name, r.range] },
  { key:'net_groups',     label:'Address Groups',   color:'#9C27B0',
    cols:['Name','Members'],
    row: r => [r.name, r.members] },
  { key:'services',       label:'Services',         color:'#F44336',
    cols:['Name','Value / Port','Type'],
    row: r => [r.name, r.value, r.type] },
  { key:'service_groups', label:'Service Groups',   color:'#E91E63',
    cols:['Name','Members'],
    row: r => [r.name, r.members] },
  { key:'zones',          label:'Zones',            color:'#00BCD4',
    cols:['Name'],
    row: r => [r.name] },
  { key:'policies',       label:'Policies',         color:'#607D8B',
    cols:['Name','Source Network','Destination Network','Service','Action','Log','Status'],
    row: r => [r.name, r.src_net, r.dst_net, r.dst_port, r.action, r.log, r.status] },
];
const PAGE = 100;

// ═══ State ═══════════════════════════════════════════════
let allData = {}, curKey = 'hosts', curPage = 1, filteredRows = [];
let sortCol = -1, sortDir = 1, searchQ = '';

// ═══ Drag & Drop ═════════════════════════════════════════
const dz = document.getElementById('drop-zone');
dz.addEventListener('dragover',  e => { e.preventDefault(); dz.classList.add('drag'); });
dz.addEventListener('dragleave', ()  => dz.classList.remove('drag'));
dz.addEventListener('drop', e => { e.preventDefault(); dz.classList.remove('drag'); handleFile(e.dataTransfer.files[0]); });

// ═══ Upload ══════════════════════════════════════════════
function handleFile(file) {
  if (!file || !file.name.endsWith('.xml')) { showToast('File must be .xml', '#ef4444'); return; }
  const fd = new FormData(); fd.append('file', file);
  document.getElementById('upload-section').style.display = 'none';
  document.getElementById('progress-section').style.display = 'block';
  animProg();
  fetch('/parse', { method:'POST', body:fd })
    .then(r => r.json())
    .then(d => {
      clearInterval(window._piv);
      document.getElementById('pbar').style.width = '100%';
      if (d.error) { showToast(d.error,'#ef4444'); resetApp(); return; }
      allData = d.data;
      renderSummary();
      switchSheet('hosts');
      document.getElementById('progress-section').style.display = 'none';
      document.getElementById('results-section').style.display = 'block';
      const total = Object.values(d.counts).reduce((a,b)=>a+b,0);
      document.getElementById('hdr-right').innerHTML =
        `<span style="font-family:var(--mono);font-size:11px;color:var(--grn)">✔ ${total.toLocaleString()} objects loaded</span>`;
    })
    .catch(e => { showToast('Error: '+e,'#ef4444'); resetApp(); });
}
function animProg() {
  const bar = document.getElementById('pbar'), lbl = document.getElementById('plabel');
  const steps=['reading xml...','parsing elements...','extracting hosts & networks...','processing services...','building zones & policies...'];
  let w=0, si=0;
  window._piv = setInterval(() => {
    w = Math.min(w + Math.random()*10, 88);
    bar.style.width = w+'%';
    if (si<steps.length) lbl.textContent = steps[si++];
  }, 400);
}

// ═══ Summary Cards ═══════════════════════════════════════
function renderSummary() {
  document.getElementById('summary-strip').innerHTML = SHEETS.map(s => {
    const cnt = (allData[s.key]||[]).length;
    return `<div class="scard" id="sc-${s.key}" onclick="switchSheet('${s.key}')">
      <div class="scard-dot" style="background:${s.color}"></div>
      <div class="scard-info">
        <div class="val" style="color:${s.color}">${cnt.toLocaleString()}</div>
        <div class="lbl">${s.label}</div>
      </div>
    </div>`;
  }).join('');
}

// ═══ Sheet Switch ═════════════════════════════════════════
function switchSheet(key) {
  curKey = key; curPage = 1; sortCol = -1; sortDir = 1;
  searchQ = document.getElementById('search-box').value = '';
  document.querySelectorAll('.scard').forEach(c => c.classList.remove('active'));
  const sc = document.getElementById('sc-'+key);
  if (sc) sc.classList.add('active');
  buildFilteredRows();
  renderHeader();
  renderPage();
}

// ═══ Build Filtered Rows ═════════════════════════════════
function buildFilteredRows() {
  const sheet = SHEETS.find(s => s.key === curKey);
  const raw   = (allData[curKey]||[]).map(r => sheet.row(r));
  filteredRows = searchQ
    ? raw.filter(r => r.some(v => String(v||'').toLowerCase().includes(searchQ)))
    : raw;
}

// ═══ Search ══════════════════════════════════════════════
function onSearch() {
  searchQ = document.getElementById('search-box').value.toLowerCase();
  curPage = 1;
  buildFilteredRows();
  renderPage();
}

// ═══ Sort ════════════════════════════════════════════════
function doSort(ci) {
  if (sortCol === ci) sortDir *= -1; else { sortCol = ci; sortDir = 1; }
  filteredRows.sort((a,b) => String(a[ci]||'').localeCompare(String(b[ci]||'')) * sortDir);
  curPage = 1; renderPage();
  // update header arrows
  document.querySelectorAll('thead th').forEach((th, i) => {
    th.classList.remove('asc','desc');
    if (i === ci) th.classList.add(sortDir===1?'asc':'desc');
  });
}

// ═══ Render Header ════════════════════════════════════════
function renderHeader() {
  const sheet = SHEETS.find(s => s.key === curKey);
  document.getElementById('thead').innerHTML =
    '<tr>' + sheet.cols.map((c,i) =>
      `<th onclick="doSort(${i})" style="color:${sheet.color}">${c}<span class="sort-ico"></span></th>`
    ).join('') + '</tr>';
}

// ═══ Render Page ═════════════════════════════════════════
function renderPage() {
  const sheet = SHEETS.find(s => s.key === curKey);
  const tot = filteredRows.length;
  const start = (curPage-1)*PAGE, end = Math.min(start+PAGE, tot);
  const wrapCols = new Set([1, 3, 4]); // Members & network cols wrap

  if (!tot) {
    document.getElementById('tbody').innerHTML =
      `<tr><td colspan="${sheet.cols.length}" class="empty-state">No records found</td></tr>`;
  } else {
    document.getElementById('tbody').innerHTML = filteredRows.slice(start, end).map(row =>
      '<tr>' + row.map((v, ci) => {
        let val = v || '';
        let cls = wrapCols.has(ci) ? 'wrap-cell' : 'nowrap';
        // Badges
        if (sheet.key === 'services' && ci === 2) {
          const bc = val==='TCP'?'badge-tcp':val==='UDP'?'badge-udp':'badge-icmp';
          val = `<span class="badge ${bc}">${val}</span>`;
        }
        if (sheet.key === 'policies' && ci === 4) {
          const bc = val==='allow'?'badge-en':val===''?'':'badge-dis';
          if(bc) val = `<span class="badge ${bc}">${val}</span>`;
        }
        if (sheet.key === 'policies' && ci === 6) {
          const bc = val==='Disabled'?'badge-dis':'badge-en';
          val = `<span class="badge ${bc}">${val}</span>`;
        }
        return `<td class="${cls}">${val}</td>`;
      }).join('') + '</tr>'
    ).join('');
  }

  // Pagination
  const tp = Math.ceil(tot/PAGE);
  document.getElementById('pg-info').textContent =
    tot ? `${(start+1).toLocaleString()}–${end.toLocaleString()} of ${tot.toLocaleString()} records` : '0 records';
  document.getElementById('btn-prev').disabled = curPage <= 1;
  document.getElementById('btn-next').disabled = curPage >= tp;
  let nums='';
  for (let p=Math.max(1,curPage-2); p<=Math.min(tp,curPage+2); p++)
    nums += `<button class="pg-btn ${p===curPage?'active':''}" onclick="gotoPage(${p})">${p}</button>`;
  document.getElementById('pg-nums').innerHTML = nums;
}

function changePage(d) { gotoPage(curPage+d); }
function gotoPage(p) {
  curPage = Math.max(1, Math.min(p, Math.ceil(filteredRows.length/PAGE)));
  renderPage();
}

// ═══ Export CSV ══════════════════════════════════════════
function exportCSV() {
  const sheet = SHEETS.find(s => s.key === curKey);
  let csv = sheet.cols.join(',') + '\n';
  filteredRows.forEach(row =>
    csv += row.map(v => '"' + String(v||'').replace(/"/g,'""') + '"').join(',') + '\n');
  const a = document.createElement('a');
  a.href = URL.createObjectURL(new Blob([csv], {type:'text/csv;charset=utf-8;'}));
  a.download = curKey + '.csv'; a.click();
  showToast('CSV exported');
}

// ═══ Export Excel ════════════════════════════════════════
function downloadExcel() {
  const btn = document.getElementById('btn-xl');
  btn.disabled = true; btn.textContent = '⏳ Generating...';
  fetch('/export_excel', {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(allData)
  })
  .then(r => r.blob())
  .then(blob => {
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob); a.download='forcepoint_export.xlsx'; a.click();
    btn.disabled=false; btn.innerHTML='⬇ Excel (All Sheets)';
    showToast('Excel downloaded!');
  })
  .catch(e => { showToast('Error: '+e,'#ef4444'); btn.disabled=false; btn.innerHTML='⬇ Excel (All Sheets)'; });
}

// ═══ Reset ═══════════════════════════════════════════════
function resetApp() {
  allData = {}; curKey='hosts'; curPage=1; filteredRows=[];
  document.getElementById('results-section').style.display='none';
  document.getElementById('progress-section').style.display='none';
  document.getElementById('upload-section').style.display='block';
  document.getElementById('hdr-right').innerHTML='';
  document.getElementById('file-input').value='';
  document.getElementById('pbar').style.width='0%';
}

// ═══ Toast ═══════════════════════════════════════════════
function showToast(msg, color) {
  const t = document.getElementById('toast');
  t.textContent=msg; t.style.background=color||'#10b981';
  t.classList.add('show'); setTimeout(()=>t.classList.remove('show'),3000);
}
</script>
</body>
</html>"""


# ═══════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════
@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/parse', methods=['POST'])
def parse():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xml'):
        return jsonify({'error': 'File must be .xml'}), 400
    try:
        content = f.read()
        data = extract_objects(content)
        counts = {k: len(v) for k, v in data.items()}
        return jsonify({'data': data, 'counts': counts})
    except ET.ParseError as e:
        return jsonify({'error': f'Invalid XML: {e}'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export_excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        buf  = build_excel(data)
        return send_file(buf,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='forcepoint_export.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print("\n" + "="*52)
    print("  Forcepoint XML Parser — Web Viewer + Excel")
    print("  http://localhost:5000")
    print("  pip install flask openpyxl")
    print("="*52 + "\n")
    app.run(debug=False, host='0.0.0.0', port=5000)